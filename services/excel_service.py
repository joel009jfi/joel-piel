import io
from datetime import datetime
from flask import make_response
import openpyxl


def exportar_ventas_excel(ventas, productos_por_pedido=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    headers = ["ID Pedido", "Fecha", "Cliente", "Email", "Producto", "Cantidad", "Precio Unit.", "Subtotal", "Total Pedido", "Estado", "Método Pago"]
    ws.append(headers)
    fila = 2
    for v in ventas:
        fecha_str = v["fecha"].strftime('%d/%m/%Y %H:%M') if hasattr(v["fecha"], 'strftime') else v["fecha"]
        total_pedido = float(v["total"]) if v["total"] else 0
        productos = (productos_por_pedido or {}).get(v["Id_pedido"], [])
        if productos:
            for i, prod in enumerate(productos):
                subtotal = float(prod.get("subtotal", 0) or (prod["cantidad"] * prod["precio_unitario"]))
                ws.cell(row=fila, column=7, value=float(prod["precio_unitario"])).number_format = '#,##0'
                ws.cell(row=fila, column=8, value=subtotal).number_format = '#,##0'
                if i == 0:
                    ws.cell(row=fila, column=1, value=v["Id_pedido"])
                    ws.cell(row=fila, column=2, value=fecha_str)
                    ws.cell(row=fila, column=3, value=v["cliente"])
                    ws.cell(row=fila, column=4, value=v["email"])
                    ws.cell(row=fila, column=9, value=total_pedido).number_format = '#,##0'
                    ws.cell(row=fila, column=10, value=v["estado"])
                    ws.cell(row=fila, column=11, value=v.get("metodo_pago", ""))
                else:
                    ws.cell(row=fila, column=1, value="")
                    ws.cell(row=fila, column=9, value="")
                ws.cell(row=fila, column=5, value=prod["nombre"])
                ws.cell(row=fila, column=6, value=prod["cantidad"])
                fila += 1
        else:
            ws.cell(row=fila, column=1, value=v["Id_pedido"])
            ws.cell(row=fila, column=2, value=fecha_str)
            ws.cell(row=fila, column=3, value=v["cliente"])
            ws.cell(row=fila, column=4, value=v["email"])
            ws.cell(row=fila, column=9, value=total_pedido).number_format = '#,##0'
            ws.cell(row=fila, column=10, value=v["estado"])
            ws.cell(row=fila, column=11, value=v.get("metodo_pago", ""))
            fila += 1
    fila += 1
    ws.cell(row=fila, column=8, value="TOTAL")
    ws.cell(row=fila, column=8).font = openpyxl.styles.Font(bold=True)
    formula_range = f"I2:I{fila-2}"
    ws.cell(row=fila, column=9).value = f"=SUM({formula_range})"
    ws.cell(row=fila, column=9).number_format = '#,##0'
    ws.cell(row=fila, column=9).font = openpyxl.styles.Font(bold=True)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    meses = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    mes_actual = meses[datetime.now().month]
    response.headers['Content-Disposition'] = f'attachment; filename=ventas joel piel mes {mes_actual}.xlsx'
    return response


def generar_excel_pedidos():
    from db import conectar, obtener_cursor
    db = conectar()
    if not db:
        return "Error de conexión", 500
    cursor = obtener_cursor(db, diccionario=True)
    cursor.execute("""
        SELECT p.Id_pedido, p.total, p.estado, p.fecha, p.metodo_pago,
               u.nombre as cliente, u.email
        FROM pedidos p
        LEFT JOIN usuarios u ON p.Id_usuario = u.Id_usuario
        ORDER BY p.fecha DESC
    """)
    ventas = cursor.fetchall()
    productos_por_pedido = {}
    for v in ventas:
        cursor.execute("""
            SELECT dp.cantidad, dp.precio_unitario, pr.nombre
            FROM detalle_pedido dp
            JOIN productos pr ON dp.Id_producto = pr.Id_producto
            WHERE dp.Id_pedido = %s
        """, (v["Id_pedido"],))
        productos_por_pedido[v["Id_pedido"]] = [
            {**p, "subtotal": p["cantidad"] * p["precio_unitario"]} for p in cursor.fetchall()
        ]
    db.close()
    return exportar_ventas_excel(ventas, productos_por_pedido)
